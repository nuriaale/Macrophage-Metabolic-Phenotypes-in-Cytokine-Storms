#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 17:35:38 2023

@author: nuria
"""

from general_functions import *

import general_functions

import numpy as np

def run_qMTA(target_model,reference_model,gene_fname,vref_dict={},gene_parameters={},gene_weight=0.5,met_weight=0.5,kpc_weight=0.5,unchanged_reaction_weight=0.5,reaction_pathway_dict={},key="Invivo_T3",coef_precision=7,max_reactions4gene=10,use_only_first_pathway=False,metabolomics_file=None,met_parameters={"target_condition":"treatment1","source_condition":"control"},kpc_file=None,kpc_parameters={},output_omit_reactions_with_more_than_max_genes=False,output_signficant_genes_only=False,normalize_by_scale_genes=True,min_flux4weight=1e-6,normalize_by_scale_unchanged_reactions=True,differential_expression_sheet_dict={},min_flux_fold_change=1e-9,qpmethod=1,n_threads=0,sample_name="",debug_prefix="",non_gene_met_reaction_weight_dict={},detailed_output=True):
    
    biocrates_name_dict={u'creatinine': u'crtn_c',
     u'spermidine': u'spmd_c',
     u'spermine': u'sprm_c',
     "creatinine":"creat_c",
     "kynurenine":"Lkynr_c",
     "taurine":"taur_c",
     "dopamine":"dopa_c",
     "dopa":"34dhphe_c",
     "alpha-aaa":"L2aadp_c",
     "histamine":"hista_c",
     "t4-oh-pro":"4hpro_LT_c",
     "serotonin":"srtn_c",
     "putrescine":"ptrc_c",
     "kynurenine":"Lkynr_c",
     "carnosine":"carn_c",
     u'ala': u'ala__L_c',
     u'arg': u'arg__L_c',
     u'asn': u'asn__L_c',
     u'asp': u'asp__L_c',
     u'carn': u'carn_c',
     u'cit': u'citr__L_c',
     u'glu': u'glu__L_c',
     u'gln': u'gln__L_c',
     u'gly': u'gly_c',
     u'his': u'his__L_c',
     u'ile': u'ile__L_c',
     u'leu': u'leu__L_c',
     u'lys': u'lys__L_c',
     u'met': u'met__L_c',
     u'orn': u'orn_c',
     u'phe': u'phe__L_c',
     u'pro': u'pro__L_c',
     u'ser': u'ser__L_c',
     u'srtn': u'srtn_c',
     u'thr': u'thr__L_c',
     u'trp': u'trp__L_c',
     u'tyr': u'tyr__L_c',
     u'val': u'val__L_c',
    "c0" :"crn_c",
    "c10":"c10dc_c",
    "c12":"c12dc_c",
    "c14":"ttdcrn_c",
    "c16":"pmtcrn_c",
    "c16:1":"hdcecrn_c",
    "c18":"stcrn_c",
    "c18:1":"odecrn_c",
    "c2":"acrn_c",
    "c3":"pcrn_c",
    "c4":"c4crn_c",
    "c6":"c6crn_c",
    "c8":"c8crn_c",
    }
    
    ref_gene_parameters={"log2_str":"log2FoldChange","log2_factor":1,"padj_str":"padj","p_th":0.25,"log2fc_th":0,"gene_str":"NCBI.gene.ID","p_weight_formula":"-1*math.log(p_value,10)","ignore_p_value":False}
    ref_gene_parameters.update(gene_parameters)
    log2_str=ref_gene_parameters["log2_str"]
    log2_factor=ref_gene_parameters["log2_factor"]
    padj_str=ref_gene_parameters["padj_str"]
    p_th=ref_gene_parameters["p_th"]
    log2fc_th=ref_gene_parameters["log2fc_th"]
    gene_str=ref_gene_parameters["gene_str"]
    p_weight_formula=ref_gene_parameters["p_weight_formula"]
    ignore_p_value=ref_gene_parameters["ignore_p_value"]
    #print log2_factor
    ####Metabolomics
    ref_met_parameters={"biocrates_name_dict":biocrates_name_dict,"target_condition":"INVIVO_Tractament_3","source_condition":"INVIVO_Control","p_adj_th_met":0.05,"convert_to_log_met":True,"p_weight_formula_met":"-1*math.log(p_value,10)","log2_factor":1,"normalize_by_scale_met":True,"normalize_p_weight":True}
    ref_met_parameters.update(met_parameters)
    biocrates_name_dict=ref_met_parameters["biocrates_name_dict"]
    target_condition=ref_met_parameters["target_condition"]
    source_condition=ref_met_parameters["source_condition"]
    p_adj_th_met=ref_met_parameters["p_adj_th_met"]
    convert_to_log_met=ref_met_parameters["convert_to_log_met"]
    p_weight_formula_met=ref_met_parameters["p_weight_formula_met"]
    log2_factor_met=ref_met_parameters["log2_factor"]
    normalize_by_scale_mets=ref_met_parameters["normalize_by_scale_met"]
    normalize_p_weight_met=ref_met_parameters["normalize_p_weight"]
    if metabolomics_file not in ("",None):
       stat_dict=read_biocrates_metabolomics_data(metabolomics_file)
       met_sink_dict, rejected_list=add_sink_reactions_with_multiple_compartments(target_model,stat_dict,biocrates_name_compartment_dict=biocrates_name_compartment_dict,lb=None,ub=None,condition="Control",precision=7,factor=1) #This is just to make sure the reactions exists, no need to define factor or precision 
       met_sink_dict, rejected_list=add_sink_reactions_with_multiple_compartments(reference_model,stat_dict,biocrates_name_compartment_dict=biocrates_name_compartment_dict,lb=None,ub=None,condition="Control",precision=7,factor=1) #This is just to make sure the reactions exists, no need to define factor or precision 
       signficant_met_dict=statistical_difference_metabolomics(stat_dict,cond1=target_condition,cond2=source_condition,convert_to_log=convert_to_log_met, p_adj_th=p_adj_th_met,met_list=biocrates_name_dict,met_sink_dict=met_sink_dict,p_weight_formula=p_weight_formula_met,log2_factor_met=log2_factor_met,normalize_p_weight=normalize_p_weight_met)
    else: signficant_met_dict={} 
    ####KPC
    ref_kpc_parameters={"kpc_name_dict":general_functions.metabolite_ex_dict,"target_condition":"","source_condition":"","p_adj_th_kpc":0.05,"p_weight_formula_kpc":"-1*math.log(p_value,10)","factor_kpc":1,"normalize_by_scale_kpc":True,"normalize_p_weight":True}
    ref_kpc_parameters.update(kpc_parameters)
    biocrates_kpc_ex_dict=ref_kpc_parameters["kpc_name_dict"]
    target_condition_kpc=ref_kpc_parameters["target_condition"]
    source_condition_kpc=ref_kpc_parameters["source_condition"]
    p_adj_th_kpc=ref_kpc_parameters["p_adj_th_kpc"]
    p_weight_formula_kpc=ref_kpc_parameters["p_weight_formula_kpc"]
    normalize_by_scale_kpc=ref_kpc_parameters["normalize_by_scale_kpc"]
    normalize_p_weight_kpc=ref_kpc_parameters["normalize_p_weight"]
    factor_kpc=ref_kpc_parameters["factor_kpc"] #Equivalent to log2 factor
    if kpc_file not in ("",None):
       modify_model_for_seahorse(target_model,remove=False) 
       modify_model_for_seahorse(reference_model,remove=False) 
       stat_dict_kpc=read_biocrates_metabolomics_data(kpc_file)
       signficant_kpc=statistical_difference_metabolomics(stat_dict_kpc,cond1=target_condition_kpc,cond2=source_condition_kpc,convert_to_log=False, p_adj_th=p_adj_th_kpc,met_list=biocrates_kpc_ex_dict,met_sink_dict=biocrates_kpc_ex_dict,p_weight_formula=p_weight_formula_kpc,log2_factor_met=factor_kpc,normalize_p_weight=normalize_p_weight_kpc)
    else:
     signficant_kpc={}
     
    up_genes, down_genes, log2fold_change_dict,   p_value_dict ,  gene_weight_dict, gene_weight_normalized_dict,=read_gene_data(fname=gene_fname,model=reference_model,log2_str=log2_str,log2_factor=log2_factor,padj_str=padj_str,p_th=p_th,log2fc_th=log2fc_th,gene_str=gene_str,p_weight_formula=p_weight_formula,sheet_dict=differential_expression_sheet_dict,ignore_p_value=ignore_p_value)
    
    quadaratic_dict, coefficient_dict, gene_target_dict, signficant_gene_list_corrected,reactions2omit=build_weight_dicts(up_genes+down_genes,reference_model,max_reactions4gene=max_reactions4gene,gene_weight=gene_weight,non_gene_met_reaction_weight=unchanged_reaction_weight,gene_weight_dict=gene_weight_normalized_dict,non_gene_met_reaction_weight_dict=non_gene_met_reaction_weight_dict,fold_change_dict=log2fold_change_dict,vref_dict=vref_dict,max_fold_change=99999999999,normalize_by_scale_genes=normalize_by_scale_genes,normalize_by_scale_unchanged_reactions=normalize_by_scale_unchanged_reactions,normalize_by_scale_mets=normalize_by_scale_mets,min_flux4weight=min_flux4weight,genes_log2_to_lineal=True,precision=coef_precision,signficant_met_dict=signficant_met_dict,met_weight=met_weight,signficant_kpc_dict=signficant_kpc,kpc_weight=kpc_weight,normalize_by_scale_kpc=normalize_by_scale_kpc,min_flux_fold_change=min_flux_fold_change)
    
    problem=create_rMTA_problem_quadratic(target_model,quadaratic_dict=quadaratic_dict, coefficient_dict=coefficient_dict,out_name="rMTA.lp",qpmethod=qpmethod)
    problem.parameters.timelimit.set(300)
    problem.parameters.threads.set(n_threads)
    problem.parameters.emphasis.numerical.set(1)

    
    try:
      problem.solve()
      status = problem.solution.get_status_string().lower()
      if status!="optimal":
          #raw_input("Press Enter to continue...") 
          problem.parameters.qpmethod.set(1)
          try:
             print( "Trying qpmethod 1" )
             problem.solve()
             status = problem.solution.get_status_string().lower()
             if status!="optimal":
                raise Exception('solver error')
          except:
             #Reset parameters and test all methods
             problem.parameters.reset()
             problem.parameters.timelimit.set(150)
             problem.parameters.threads.set(n_threads)
             for emphasis_numerical in (0,1):
               for method in  (2,3,4,5,6,4,1):
                 try:
                     print( "Trying qpmethod",method,"emphasis",emphasis_numerical )
                     problem.parameters.emphasis.numerical.set(emphasis_numerical)
                     problem.parameters.qpmethod.set(method)
                     problem.solve()
                     status = problem.solution.get_status_string().lower()
                     print( "qpmethod",method,"emphasis",emphasis_numerical,status )
                     if status=="optimal": break
                 except:
                     print( "solver error" )
               if status=="optimal": break
      if  status!="optimal": #Tunning parameters should not do anything not already done, but we leave it just in case
     
          problem.parameters.qpmethod.set(1) 
          print( "tuning parameters" )
          try:
             problem.parameters.tune_problem()
             problem.parameters.write_file("parameters.txt")
             problem.solve()
             status = problem.solution.get_status_string().lower()
             print( "tunning",status )
             if status!="optimal":
                raise Exception('solver error')
          except:
             print( "Swicthing to default qpMethod to attempt to get a solution even if its not fully optimal"  )
             problem.parameters.qpmethod.set(0)
             problem.solve()
             if status!="optimal":
                problem.write("debug.lp") 
                f=open(debug_prefix+"debug.txt","a")
                f.write(sample_name+"\n")
                f.close()
                
      solution=cplex_format_solution(problem, target_model) 
      for reaction in reference_model.reactions:
        if reaction.id not in solution.x_dict:
           solution.x_dict[reaction.id]=0 
    except: raise Exception('solver error')
    
    if detailed_output:       
       reaction_dict_dict, variation_dict_dict , output_sheet=process_mta_results(reference_model,vref_dict=vref_dict,vres_dict=solution.x_dict,gene_target_dict=gene_target_dict,up_genes=up_genes,down_genes=down_genes,p_value_dict=p_value_dict,key=key,omit_reactions=output_omit_reactions_with_more_than_max_genes,reactions_2_omit=reactions2omit,reaction_pathway_dict=reaction_pathway_dict,use_only_first_pathway=use_only_first_pathway,signficant_met_dict=signficant_met_dict,signficant_kpc_dict=signficant_kpc,signficant_genes_only=output_signficant_genes_only)
    else:
       reaction_dict_dict={}
       variation_dict_dict={}
       output_sheet={}
    return output_sheet, solution.x_dict, reaction_dict_dict, variation_dict_dict,up_genes, down_genes, log2fold_change_dict,   p_value_dict ,  gene_weight_dict,signficant_met_dict,signficant_kpc

def statistical_difference_metabolomics(stat_dict,cond1="target",cond2="control",convert_to_log=True, p_adj_th=0.05,met_list=None,p_weight_formula="-1*math.log(p_value,10)",normalize_p_weight=True,met_sink_dict={},log2_factor_met=1):
    #log2_factor_met is used to do the reverse MTA
    if met_list==None:
       met_list=stat_dict.keys() 
    p_value_list=[]
    signficant_met_dict={}

    for met in sorted(met_list):
        if met not in stat_dict:
           continue
        if cond1 not in stat_dict[met]:
           continue 
        if cond2 not in stat_dict[met]:
           continue 
        values1=stat_dict[met][cond1]["values"]
        values2=stat_dict[met][cond2]["values"]
        mean1=stat_dict[met][cond1]["mean"]
        mean2=stat_dict[met][cond2]["mean"] 
        if convert_to_log:
           comparisson=cond1+"/"+cond2 
           values1=[math.log(x,2) for x in values1] 
           values2=[math.log(x,2) for x in values2]
        else:
           comparisson=cond1+"-"+cond2  
        from scipy import stats
        result=stats.ttest_ind(values1, values2, axis=None, equal_var=True, nan_policy='propagate')
        p_value=result[1]
        p_value_list.append(p_value)
        fc=mean1/mean2
        if convert_to_log:
           lfc= math.log(fc,2)*log2_factor_met
        else:
            lfc=None
        if log2_factor_met==1:
           local_dict={"p_value":p_value,"fc":fc,"logfc":lfc,"mean_source":mean2,"mean_target":mean1}
        elif log2_factor_met==-1:
           local_dict={"p_value":p_value,"fc":fc,"logfc":lfc,"mean_source":mean1,"mean_target":mean2}
        else:
            raise Exception('log2_factor_met should be 1 or -1')
        if "statistics" not in stat_dict:
            stat_dict[met]["statistics"]={}
        stat_dict[met]["statistics"][comparisson]=local_dict
    ##Correct p value
    from statsmodels.stats.multitest import multipletests
    rejected, corrected_p_values, alphacSidak, alphacBonf = multipletests(p_value_list, method='fdr_bh')

    n_counter=0
    for met in sorted(met_list):
         if met not in stat_dict:
            continue
         if "statistics" not in stat_dict[met]:
             continue 
         stat_dict[met]["statistics"][comparisson]["p_adj"]=corrected_p_values[n_counter]
         p_value=corrected_p_values[n_counter]
         values1=stat_dict[met][cond1]["values"] #To be used in formulas
         values2=stat_dict[met][cond2]["values"] #To be used in formulas        
         sd1=np.std(values1) #To be used in formulas
         sd2=np.std(values2) #To be used in formulas
         stat_dict[met]["statistics"][comparisson]["p_weight"]=eval(p_weight_formula)
         if  corrected_p_values[n_counter]<p_adj_th:
             signficant_met_dict[met]=stat_dict[met]["statistics"][comparisson]
             signficant_met_dict[met]["reaction"]=met_sink_dict.get(met)
         n_counter+=1
    
    if normalize_p_weight:
       p_factor=p_mean=np.mean([ signficant_met_dict[met]["p_weight"] for met in signficant_met_dict])
    else:
       p_factor=p_mean=1.0 
    p_mean_normalized_dict={met:round_sig(signficant_met_dict[met]["p_weight"]/p_mean,3) for met in signficant_met_dict}
    for met in signficant_met_dict:
        signficant_met_dict[met]["p_weight_norm"]=p_mean_normalized_dict[met]
    return signficant_met_dict

def read_gene_data(fname,model,log2_str="log2FoldChange",log2_factor=1,padj_str="padj",p_th=0.25,log2fc_th=0,gene_str="NCBI.gene.ID",p_weight_formula="-1*math.log(p_value,10)",sheet_dict={},ignore_p_value=False):
    up_genes=[]
    down_genes=[]
    log2fold_change_dict={}
    p_value_dict={}
    milp_weight_dict={}
    milp_weight_list=[] #For normalization
    if sheet_dict=={}:
       sheet_dict=read_spreadsheets(fname)
    sheet=sheet_dict[list(sheet_dict.keys())[0]]           
    gene_n=0 #If gene is not defined we will assume is the first one
    for n_row,row in enumerate(sheet): 
      if n_row==0: #Header
         for n, element in enumerate(row):
           if element==None:
               continue   
           if log2_str==element:
               log2_n=n
           elif padj_str==element:
                p_adj_n=n
           elif gene_str==element:
                gene_n=n
         continue   
         
      gene_id= str(row[gene_n])
      log2fc=row[log2_n]
      if not ignore_p_value:
         p_value=row[p_adj_n]
      else:
         p_value=0 
      if p_value in ("NA","",None):
         continue
      log2fc=float(log2fc)*log2_factor
      p_value=float(p_value)

      if gene_id in model.genes:
         #print row
         if abs(log2fc)>log2fc_th and p_value<p_th:
            if log2fc>=0:
               up_genes.append(gene_id) 
            elif log2fc<0:
               down_genes.append(gene_id)
            log2fold_change_dict[gene_id]=log2fc
            p_value_dict[gene_id]=p_value
            milp_weight=eval(p_weight_formula)
            milp_weight_dict[gene_id]=milp_weight
            milp_weight_list.append(milp_weight)    
    
    milp_mean=np.mean([milp_weight_dict[x] for x in milp_weight_dict])
    milp_weight_normalized_dict={x:round(milp_weight_dict[x]/milp_mean,3) for x in milp_weight_dict}
    return  up_genes, down_genes, log2fold_change_dict,   p_value_dict ,  milp_weight_dict, milp_weight_normalized_dict

def build_weight_dicts(signficant_gene_list,cobra_model,max_reactions4gene=10,gene_weight=0.5,met_weight=1,kpc_weight=1,non_gene_met_reaction_weight=0.5,gene_weight_dict={},non_gene_met_reaction_weight_dict={},fold_change_dict={},vref_dict={},max_fold_change=99999999999,min_flux4weight=1e-6,normalize_by_scale_genes=True,normalize_by_scale_unchanged_reactions=True,genes_log2_to_lineal=True,precision=6,signficant_met_dict={},normalize_by_scale_mets=True,signficant_kpc_dict={},normalize_by_scale_kpc=True,min_flux_fold_change=1e-9):
    gene_met_reactions=set()
    genes_to_omit=[]
    quadaratic_dict={}
    coefficient_dict={}
    gene_target_dict={}
    reactions2omit=[]
    for gene_id in signficant_gene_list:
       gene=cobra_model.genes.get_by_id(gene_id) 
       if len(gene.reactions)>max_reactions4gene:
          genes_to_omit.append(gene_id)
    
    
    signficant_gene_list_corrected=[x for x in signficant_gene_list if x not in genes_to_omit]
    ################## Signficant weight
    for gene_id in signficant_gene_list_corrected:
        if not gene_id  in cobra_model.genes:
            continue
        fold_change=max(min(max_fold_change,fold_change_dict[gene_id]),-max_fold_change)
        if genes_log2_to_lineal:
                fold_change=pow(2,fold_change)
        for reaction in cobra_model.genes.get_by_id(gene_id).reactions:
            rid=reaction.id
            gene_met_reactions.add(rid)
            vref=vref_dict[rid]
            #Todo add correction to genes catalyzing many reactions
            target_vref=round(vref*fold_change,7)
            weight=float(gene_weight)
            if normalize_by_scale_genes==True:
               flux_factor= max(abs(pow(vref-target_vref,2)),min_flux4weight)
               #Minimal fold change
               flux_factor_min= max(abs(pow(vref*0.999,2)),min_flux4weight)
   
               weight/= max(flux_factor,0)
         
            if gene_id in gene_weight_dict:
               weight*=gene_weight_dict[gene_id]
  
            if abs(vref)<min_flux_fold_change:
               weight=0 
            if rid not in coefficient_dict or rid not in quadaratic_dict:
               coefficient_dict[rid]=0
               quadaratic_dict[rid]=0
            if gene_id not in gene_target_dict:
               gene_target_dict[gene_id]={}
            coefficient_dict[rid]+=-2*target_vref*weight#-2ab*weight
            quadaratic_dict[rid]+=weight*2 #b2 Multipy by 2
            gene_target_dict[gene_id][rid]=round_sig(target_vref,4)
            if weight==0:
               gene_target_dict[gene_id][rid]=0  
    
    #Metabolomics

    for met in signficant_met_dict:
        rid=signficant_met_dict[met]["reaction"]
        if rid in cobra_model.reactions:
            try:
              fold_change=math.pow(2,signficant_met_dict[met]["logfc"])
            except:
              fold_change=signficant_met_dict[met]["fc"]  
            gene_met_reactions.add(rid)
            vref=vref_dict[rid]
            target_vref=vref*fold_change
            weight=float(met_weight)
            if normalize_by_scale_mets:
               weight/= max(abs(pow(vref-target_vref,2)),min_flux4weight)
            weight*=signficant_met_dict[met]["p_weight_norm"]
            """if weight!=0:
               weight=round_sig(weight,precision)"""
            if rid not in coefficient_dict or rid not in quadaratic_dict:
               coefficient_dict[rid]=0
               quadaratic_dict[rid]=0
            coefficient_dict[rid]+=-2*target_vref*weight#-2ab*weight
            quadaratic_dict[rid]+=weight*2 #b2 Multipy by 2
           
            signficant_met_dict[met]["target_vref"]=target_vref
    #Minimization weight for remaining reactions:
    ####KPC

    for met in signficant_kpc_dict:
        rid=signficant_kpc_dict[met]["reaction"]
        if rid in cobra_model.reactions:
           
            gene_met_reactions.add(rid)
            vref=vref_dict[rid]
            difference=signficant_kpc_dict[met]["mean_target"]-signficant_kpc_dict[met]["mean_source"] #We use this format to allow reversing
            target_vref=vref+difference
            weight=float(kpc_weight)
            if normalize_by_scale_kpc:
               weight/= max(abs(pow(vref-target_vref,2)),min_flux4weight)
            weight*=signficant_kpc_dict[met]["p_weight_norm"]
            
            if rid not in coefficient_dict or rid not in quadaratic_dict:
               coefficient_dict[rid]=0
               quadaratic_dict[rid]=0
            coefficient_dict[rid]+=-2*target_vref*weight#-2ab*weight
            quadaratic_dict[rid]+=weight*2 #b2 Multipy by 2
        
            signficant_kpc_dict[met]["target_vref"]=target_vref
    non_gene_met_reactions=[x for x in cobra_model.reactions if x.id not in  gene_met_reactions]
  
    for reaction in non_gene_met_reactions:
        genes_id=[x.id for x in reaction.genes]
        if len(set(genes_id).intersection(genes_to_omit))>0:
           reactions2omit.append(reaction.id) 
        rid=reaction.id

        if rid in vref_dict:
            vref=vref_dict[rid]
        else:
            continue
        weight=non_gene_met_reaction_weight
        if normalize_by_scale_unchanged_reactions==True:
    
           weight/= max(abs(vref),min_flux4weight) #This works better than squared

        if rid in non_gene_met_reaction_weight_dict:
           weight*=non_gene_met_reaction_weight_dict[rid]

        coefficient_dict[rid]=-2*vref*weight#-2ab*weight
        quadaratic_dict[rid]=weight*2 #b2/2 Multipy by 2 
    return quadaratic_dict, coefficient_dict, gene_target_dict, signficant_gene_list_corrected,reactions2omit


def process_mta_results(analyzed_model,vref_dict={},vres_dict={},gene_target_dict={},up_genes=[],down_genes=[],p_value_dict={},key="MTA_results",omit_reactions=False,reactions_2_omit=[],reaction_pathway_dict={},use_only_first_pathway=True,signficant_met_dict={},signficant_kpc_dict={},signficant_genes_only=False):
  if reaction_pathway_dict in ("",{},None):
     omit_pathways=True
     reaction_pathway_dict={}
  else:
     omit_pathways=False 
  reactions_without_pathway=[x.id for x in analyzed_model.reactions if (reaction_pathway_dict.get(x.id)==[] or reaction_pathway_dict.get(x.id)==None)]
  for x in reactions_without_pathway:
    if "SINK" in x:
       reaction_pathway_dict[x]=["metabolomics"] 
    elif "EX_" in x:
       reaction_pathway_dict[x]=["Exchange_reactions"] 
    else:
       reaction_pathway_dict[x]=["no pathway assigned"]

  variation_dict_dict={}
  reaction_dict_dict={}

  if True:
    variation_dict={}
    reaction_dict={}
    reaction_dict_dict[key]=reaction_dict
    variation_dict_dict[key]=variation_dict
    reaction_gene_dict={}
    
    limited_output_flag=max([len(x.genes) for x in analyzed_model.reactions])==1 #In case we are working with reactions
    for reaction in analyzed_model.reactions:
          omited_flag=False   
          reaction_id=reaction.id
          reaction_gene_dict[reaction_id]={"genes_up":[],"genes_down":[]}
          reaction_genes_up=[]
          for x in reaction.genes:
                if x.id in up_genes and not omited_flag:
                   try: 
                      if limited_output_flag:
                         reaction_str= str(gene_target_dict[x.id][reaction.id])
                      else:    
                          reaction_str=str(x.id)+"("+str(gene_target_dict[x.id][reaction.id])+"/"+str(round(p_value_dict[x.id],3))+")"
                      reaction_genes_up.append(reaction_str)
                      reaction_gene_dict[reaction_id]["genes_up"].append(x.id)
                   except:
                      pass     
          reaction_genes_down=[]
          for x in reaction.genes:
                if x.id in down_genes and not omited_flag:
                   try:
                     if limited_output_flag:
                         reaction_str= str(gene_target_dict[x.id][reaction.id])             
                     else:
                         reaction_str=str(x.id)+"("+str(gene_target_dict[x.id][reaction.id])+"/"+str(round(p_value_dict[x.id],3))+")"
                     reaction_genes_down.append(reaction_str)
                     reaction_gene_dict[reaction_id]["genes_down"].append(x.id)
                   except:
                     pass                     
          signficant_genes_flag=len(reaction_genes_up+reaction_genes_down)>0 or not signficant_genes_only
          for n,system in enumerate(reaction_pathway_dict[reaction_id]):
            if use_only_first_pathway and n>0:
                  break
            if system not in variation_dict:
               variation_dict[system]={"n_increase":0, "n_decrease":0,"increased":[],"decreased":[],"total_increase":0,"total_increase_genes":0,"total_decrease_genes":0,"total_decrease":0,"omitted":[],"n_omitted":0,"total_increase_normalized":0,"total_decrease_normalized":0,"normalized_n_reaction_score_increased":0,"normalized_n_reaction_score_decreased":0,"total_flux_vref":0,"total_flux_vres":0,"log2fc":0,"reaction_fc":[],"reaction_log2fc":[],"genes_up":[],"genes_down":[]} 
            if reaction_id in vref_dict:
                vref=vref_dict[reaction_id]
            else:
                vref=0
            vres=vres_dict[reaction_id]
            variation_dict[system]["genes_up"]+=reaction_gene_dict[reaction_id]["genes_up"]
            variation_dict[system]["genes_down"]+=reaction_gene_dict[reaction_id]["genes_down"]
            if signficant_genes_flag and not (reaction_id in reactions_2_omit and omit_reactions):
               variation_dict[system]["total_flux_vref"]+=abs(vref)
               variation_dict[system]["total_flux_vres"]+=abs(vres)
        
            if reaction_id in reactions_2_omit and omit_reactions:
               #print "omited " +reaction_id
               variation_dict[system]["n_omitted"]+=1
               variation_dict[system]["omitted"].append(reaction_id)
               omited_flag=True
               diff="omited"
     
            else:
               diff=vres-vref
               omited_flag=False 
               
            if (abs(vres)-abs(vref))>=1e-6 and not omited_flag:
               variation_dict[system]["n_increase"]+=1
          
               variation_dict[system]["increased"].append(reaction_id)
               diff=abs(diff)
               if signficant_genes_flag:
                  variation_dict[system]["total_increase"]+= (abs(vres)-abs(vref))
      
            elif abs(vres)-abs(vref)<=-1e-6 and not omited_flag:
               variation_dict[system]["n_decrease"]+=1

               variation_dict[system]["decreased"].append(reaction_id)
               if signficant_genes_flag:
                  variation_dict[system]["total_decrease"]+= abs(abs(vres)-abs(vref))
     
            if abs(vref)>1e-5:
               reaction_fc=(max(abs(vres),1e-5)/abs(vref))
      
               reaction_log2fc=math.log(reaction_fc,2)
               if signficant_genes_flag:
                  variation_dict[system]["reaction_fc"].append(reaction_fc)
                  variation_dict[system]["reaction_log2fc"].append(reaction_log2fc)
            else:
               reaction_log2fc=""
            try:
               reaction_dict[reaction_id]=[vref,vres,diff,reaction_log2fc,float((reaction_genes_up+reaction_genes_down)[0])] 
            except:
               reaction_dict[reaction_id]=[vref,vres,diff,reaction_log2fc,",".join(reaction_genes_up),",".join(reaction_genes_down)]
  ###Metabolomics: 
  for met in signficant_met_dict:
      reaction=signficant_met_dict[met]["reaction"]
      p_value=str(round(signficant_met_dict[met]['p_adj'],6))
      lfc=str(signficant_met_dict[met]['logfc'])
    
      target_vres=str(round(signficant_met_dict[met]["target_vref"],7))
      metabolomics_str="logfc:"+lfc+" p_adj:"+str(p_value)+" target_vres:"+target_vres
      reaction_dict[reaction][-2]=metabolomics_str
  for met in signficant_kpc_dict:
      reaction=signficant_kpc_dict[met]["reaction"]
      p_value=str(round(signficant_kpc_dict[met]['p_adj'],6))
      lfc=str(signficant_kpc_dict[met]['logfc'])
 
      target_vres=str(round(signficant_kpc_dict[met]["target_vref"],7))
      metabolomics_str="logfc:"+lfc+" p_adj:"+str(p_value)+" target_vres:"+target_vres
      reaction_dict[reaction][-2]=metabolomics_str      
  output_sheet={}
  model=analyzed_model
  header_2=["rid","name","reaction","genes","vref","vres","variation","log2fc","reaction_genes_up","reaction_genes_down"]
  header1=["system","n_omited","omited","n_genes_up","n_genes_down","total_flux_vref","total_flux_vres","n_increase","n_decrease","increased","decreased","total_increase","total_decrease","total_variation","total_fold_change","Log2_fold_change","Average log2FC"]
  for key in variation_dict_dict:
    variation_dict= variation_dict_dict[key]
    reaction_dict=reaction_dict_dict[key]
    output_sheet[key+"_a"]=[header1]
    output_sheet[key+"_b"]=[]
    
    for system in variation_dict:
      if system=="no pathway assigned":# or (variation_dict[system]["total_increase"]-variation_dict[system]["total_decrease"])==0:
         continue
      fc=variation_dict[system]["total_flux_vres"]/max(variation_dict[system]["total_flux_vref"],1e-6)
      if fc!=0:
         logfc=math.log(fc,2)
      else:
         logfc=0
      if len(variation_dict[system]["reaction_log2fc"])>0:
         mean_log2=np.mean(variation_dict[system]["reaction_log2fc"])
      else:
         mean_log2=0    
      row=[system,variation_dict[system]["n_omitted"],str(variation_dict[system]["omitted"]),len(set(variation_dict[system]["genes_up"])),len(set(variation_dict[system]["genes_down"])),variation_dict[system]["total_flux_vref"],variation_dict[system]["total_flux_vres"],variation_dict[system]["n_increase"],variation_dict[system]["n_decrease"],str(variation_dict[system]["increased"]),str(variation_dict[system]["decreased"]),variation_dict[system]["total_increase"],variation_dict[system]["total_decrease"],variation_dict[system]["total_increase"]-variation_dict[system]["total_decrease"],fc,logfc,mean_log2]
      output_sheet[key+"_a"].append(row)
      output_sheet[key+"_b"].append(header1)
      output_sheet[key+"_b"].append(row)
      output_sheet[key+"_b"].append(["","increased"]+header_2)
      for reaction_id in variation_dict[system]["increased"]:

        reaction=model.reactions.get_by_id(reaction_id)
        reaction_row=["","",reaction.id,reaction.name,reaction.reaction,reaction.gene_reaction_rule]+reaction_dict[reaction_id]
        output_sheet[key+"_b"].append(reaction_row)
      output_sheet[key+"_b"].append(["","decreased"]+header_2)
      for reaction_id in variation_dict[system]["decreased"]:
   
          reaction=model.reactions.get_by_id(reaction_id)
          reaction_row=["","",reaction.id,reaction.name,reaction.reaction,reaction.gene_reaction_rule]+reaction_dict[reaction_id]
          output_sheet[key+"_b"].append(reaction_row)
    
  
    if omit_pathways:
        output_sheet={} 
    output_sheet[key+"_all_reactions"]=[["Reaction ID","Reaction Name","Reaction id","Reaction","Vref","Vres","Variation","Log2FC","Target flux"]]
    for reaction in analyzed_model.reactions:
        reaction_row=[reaction.id,reaction.name,reaction.reaction,reaction.gene_reaction_rule]
        if reaction.id in reaction_dict:
           reaction_row+=reaction_dict[reaction.id]  
        output_sheet[key+"_all_reactions"].append(reaction_row)
    
    return reaction_dict_dict, variation_dict_dict , output_sheet

def read_biocrates_metabolomics_data(metabolomics_file="",max_NA_fraction=0.5):
    #Metaboanlyst like, Assume samples in rows and the fisrt row is samples name followes by condition
    data_sheet=read_spreadsheets(metabolomics_file)
    data_dict={}
    n_met_dict={}
    for sheet in data_sheet:
      na_counter={}  
      max_na=max_NA_fraction*(len(data_sheet[sheet])-1) #This only for the fist sheet   
      for n_row,row in enumerate(data_sheet[sheet]):

       condition=row[1]
       sample=row[0]
       for n_col,col in enumerate(row):
           if n_col>1:
              if n_row==0:
                 n_met_dict[n_col]=col.lower()
  
              else:
                 if condition=="":
                    continue
                 met=n_met_dict[n_col] 

                 try:
                   value=float(col)
                 except:
     
                     if met not in na_counter:
                        na_counter[met]=1
                     else:
                        na_counter[met]+=1
                     
                     value="Na"   
                 if value=="Na":
                    continue 
                 if met not in data_dict:
                    data_dict[met]={}

                 if condition not in data_dict[met]:
                    data_dict[met][condition]=[]
                 data_dict[met][condition].append(value)

    
      #Remove NA
      for met in na_counter:
        n_na=na_counter[met]
        if n_na>=max_na and met in data_dict:
           del(data_dict[met]) 
    stat_dict={}
    for met in data_dict:
           met=met.lower()
           stat_dict[met]={}
           for condition in data_dict[met]:
               stat_dict[met][condition]={}
               values=data_dict[met][condition]
               mean=np.mean(values)
               std=np.std(values)
               stat_dict[met][condition]["mean"]=mean
               stat_dict[met][condition]["std"]=std
               stat_dict[met][condition]["se"]=std/float(len(values))
               stat_dict[met][condition]["values"]=values
    return stat_dict

import openpyxl 

import string
alphabet=list(string.ascii_uppercase)
alphabet2=[]
for letter1 in alphabet:
    for letter2 in alphabet:
        alphabet2.append(letter1+letter2) 

alphabet+=alphabet2
from openpyxl import load_workbook

def read_spreadsheets(file_names=None,csv_delimiter=',',more_than_1=True,tkinter_title="Chose a file"):
    if not isinstance(file_names,list):
       file_names=[file_names]
    condition_rows_dict={}
    for file_name in file_names:
        file_type=file_name[-6:] #Take some extra caracters just in case
        if any(x in file_type for x in [".xlsx",".xlsm",".xltx",".xltm"]):
           print( file_type )
           wb = load_workbook(file_name, read_only=True,data_only=True)
           for ws in wb.worksheets:
                  condition=ws.title
                  #
                  if "Sheet" in condition:
                     condition=condition.replace("Sheet","Label_")
                  elif "Hoja" in condition:
                     condition=condition.replace("Hoja","Label_") 
                  condition_rows_dict[condition]=[]
                  for xlsx_row in ws.rows:
                      row=[]
                      for cell in xlsx_row:
                          row.append(cell.value)
                      condition_rows_dict[condition].append(row)
        else:
           csv_file=open(file_name)
           csv_reader=csv.reader(csv_file,delimiter=csv_delimiter)
           row_list=list(csv_reader)
           condition=(file_name.split(".")[0]).split("/")[-1] #Get the name of the file minus the extension and the path
           condition_rows_dict[condition]=row_list
           csv_file.close() 
    return condition_rows_dict


def write_spreadsheet(file_name,sheet_row_data_dict,sheet_order=None,force_extension=False):
    if sheet_order==None or sheet_order==[]:
       sheet_order=sorted(sheet_row_data_dict)
    print( file_name )
    if ".xlsx" in file_name:
       wb = openpyxl.Workbook()
       for n_sheet,sheet_id in enumerate(sheet_order):
           sheet_title = (str(n_sheet)+"_"+sheet_id[:27] + '..') if len(sheet_id) > 31 else sheet_id
           if n_sheet==0:
              sheet = wb.active
              sheet.title=sheet_title
           else:
                          
              sheet=wb.create_sheet(title=sheet_title)
           for n_row, row in enumerate(sheet_row_data_dict[sheet_id]):
               for n_col,data in enumerate(sheet_row_data_dict[sheet_id][n_row]):
                    try:
                      sheet[alphabet[n_col]+str(n_row+1)]=data
                    except:
                      sheet[alphabet[n_col]+str(n_row+1)]="encoding_error" 
       wb.save(file_name)
       return
    else:
        if ".txt" not in file_name.lower() and ".csv" not in file_name.lower() and force_extension: 
           file_name+=".csv" 
        csvFile = open(file_name, 'w')
        outputWriter = csv.writer(csvFile)
        for n_sheet,sheet_id in enumerate(sheet_order):
            if len(sheet_order)>1:
               outputWriter.writerow(['###'+sheet_id+"###"])
            for n_row, row in enumerate(sheet_row_data_dict[sheet_id]):
                outputWriter.writerow(row)
            if len(sheet_order)>1: 
                outputWriter.writerow(["######"])
        csvFile.close()
        return       